import openpyxl
import time
import pyautogui
from pages.login_page import loginPage
from pages.home_page import HomePage
from pages.cart_page import CartPage
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoAlertPresentException
from pages.checkout_page import CheckoutPage
import pytest
@pytest.mark.order(2)
def test_cart_checkout_with_excel(driver):
    ...

def test_cart_checkout_flow(driver):
    def close_alert_popup():
        time.sleep(2)  # Wait for popup to appear
        pyautogui.press('enter')  # Simulates pressing Enter
        print("Popup closed by pressing Enter.")
    driver.get("https://www.saucedemo.com/")
    close_alert_popup()

    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.ID, "user-name")))

    # Login
    wb = openpyxl.load_workbook('testdata/credentials.xlsx')
    sheet = wb.active
    username = sheet.cell(row=2, column=1).value
    password = sheet.cell(row=2, column=2).value

    login = loginPage(driver)
    print("Username field present:", driver.find_element(By.ID, "user-name").is_displayed())
    login.enter_username(username)
    time.sleep(2)
    login.enter_password(password)
    time.sleep(2)
    login.click_login_button()
    time.sleep(2)


    # Add to Cart
    home = HomePage(driver)
    time.sleep(2)
    home.add_to_cart()
    time.sleep(2)
    home.go_to_cart()

    # Remove from Cart and Proceed
    cart = CartPage(driver)
    cart.remove_item()
    time.sleep(2)
    cart.proceed_to_checkout()

    # Checkout Flow
    checkout = CheckoutPage(driver)
    checkout.enter_details("Chintu", "Tester", "500001")
    time.sleep(2)
    checkout.click_continue()
    time.sleep(2)
    checkout.click_finish()
    time.sleep(2)

    assert checkout.get_confirmation() == "Thank you for your order!"
