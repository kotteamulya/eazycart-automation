import openpyxl
import pytest
from selenium.common import NoAlertPresentException
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from pages.login_page import loginPage

@pytest.mark.order(3)
def test_invalid_login_with_excel(driver):
    ...
def test_invalid_login_with_excel(driver):
    driver.get("https://www.saucedemo.com/")
    # Load Excel file
    workbook = openpyxl.load_workbook('C:/eazycart_automation/testdata/credentials.xlsx')
    sheet = workbook.active

     # Read username and password from Excel (row 3 = invalid user)
    username = sheet.cell(row=3, column=1).value
    password = sheet.cell(row=3, column=2).value

        # Use Login Page methods
    login_page = loginPage(driver)
    login_page.enter_username(username)
    login_page.enter_password(password)
    login_page.click_login_button()

        # Verify error message
    error = WebDriverWait(driver, 15).until(
      EC.visibility_of_element_located((By.XPATH, "//h3[@data-test='error']"))
     )
    assert 'Epic sadface' in error.text