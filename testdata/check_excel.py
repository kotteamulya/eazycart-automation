import openpyxl

wb = openpyxl.load_workbook("C:/eazycart_automation/testdata/credentials.xlsx")
sheet = wb.active
username = sheet.cell(row=2, column=1).value
password = sheet.cell(row=2, column=2).value

print("Username:", username)
print("Password:", password)
